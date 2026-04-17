from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.prompt import Prompt, FloatPrompt
from rich.text import Text
from rich import box

console = Console()

TOMATO_ART = r"""
    ⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢰⣄⠀⠀⠀⠀⠀⠀⠀
    ⠀⠀⠀⠀⠀⠀⠻⡶⣶⣆⣸⣸⣿⣃⣀⡀⣀⣀⠀⠀⠀⠀
    ⠀⠀⢠⠟⠋⠉⠨⣽⠯⠛⣛⠯⠺⡻⣿⣽⣒⢂⠀⠀⠀⠀
    ⠀⡨⠏⠀⠀⠀⠈⠀⠀⠀⡏⠁⠀⠀⠉⠓⠭⠉⠑⢦⠀⠀
    ⢴⠃⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⣅
    ⣸⠅⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠌
    ⠺⡄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡸
    ⠀⠛⢦⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⡞⠀
    ⠀⠀⠀⠙⠙⠮⠦⠴⠤⢤⣤⣤⣀⣤⡤⠒⠒⠉⠁⠀
""".strip().replace("⠀", " ")


class DynamicBillSplitter:
    def __init__(self):
        self.net_balances: dict[str, float] = {}
        self.history: list[dict] = []

    def add_bill(self, amount: float, payer: str, consumers: list[str]) -> None:
        if payer not in self.net_balances:
            self.net_balances[payer] = 0.0
        for person in consumers:
            if person not in self.net_balances:
                self.net_balances[person] = 0.0

        split_amount = amount / len(consumers)
        self.net_balances[payer] += amount
        for person in consumers:
            self.net_balances[person] -= split_amount

        self.history.append({
            "id": len(self.history) + 1,
            "payer": payer,
            "amount": amount,
            "consumers": consumers,
        })

    def delete_bill(self, bill_id: int) -> bool:
        idx = bill_id - 1
        if idx < 0 or idx >= len(self.history):
            return False
        entry = self.history[idx]
        split_amount = entry["amount"] / len(entry["consumers"])
        self.net_balances[entry["payer"]] -= entry["amount"]
        for person in entry["consumers"]:
            self.net_balances[person] += split_amount
        self.history.pop(idx)
        for i in range(idx, len(self.history)):
            self.history[i]["id"] = i + 1
        return True

    def get_settlements(self) -> list[tuple[str, str, float]]:
        debtors = []
        creditors = []
        for person, balance in self.net_balances.items():
            if balance < -0.01:
                debtors.append([person, abs(balance)])
            elif balance > 0.01:
                creditors.append([person, balance])

        settlements = []
        d_idx, c_idx = 0, 0
        while d_idx < len(debtors) and c_idx < len(creditors):
            debtor_name, d_amt = debtors[d_idx]
            creditor_name, c_amt = creditors[c_idx]
            payment = min(d_amt, c_amt)
            settlements.append((debtor_name, creditor_name, payment))
            debtors[d_idx][1] -= payment
            creditors[c_idx][1] -= payment
            if debtors[d_idx][1] < 0.01:
                d_idx += 1
            if creditors[c_idx][1] < 0.01:
                c_idx += 1
        return settlements

    @property
    def known_people(self) -> list[str]:
        return list(self.net_balances.keys())


def show_banner():
    console.print(f"[red]{TOMATO_ART}[/red]")
    console.print(
        Panel(
            "[bold red]TOMATO BILL SPLITTER[/bold red]",
            border_style="bright_red",
            box=box.DOUBLE,
            expand=False,
            padding=(0, 6),
        )
    )


def show_menu():
    menu = Table(show_header=False, box=None, padding=(0, 2))
    menu.add_column(style="bold cyan", justify="center")
    menu.add_column(style="white")
    menu.add_row("1", "Add a bill")
    menu.add_row("2", "View balances")
    menu.add_row("3", "View bill history")
    menu.add_row("4", "Delete a bill")
    menu.add_row("5", "Settle up & exit")
    console.print()
    console.print(Panel(menu, title="[bold]Menu[/bold]", border_style="cyan", expand=False))


def show_balances(splitter: DynamicBillSplitter):
    if not splitter.net_balances:
        console.print("[dim]No participants yet.[/dim]")
        return
    table = Table(
        title="Current Balances",
        box=box.ROUNDED,
        border_style="blue",
        title_style="bold blue",
    )
    table.add_column("Person", style="bold")
    table.add_column("Balance", justify="right")
    for person, balance in sorted(splitter.net_balances.items()):
        if balance > 0.01:
            style = "green"
            display = f"+{balance:.2f}"
        elif balance < -0.01:
            style = "red"
            display = f"{balance:.2f}"
        else:
            style = "dim"
            display = "0.00"
        table.add_row(person, Text(display, style=style))
    console.print(table)


def show_history(splitter: DynamicBillSplitter):
    if not splitter.history:
        console.print("[dim]No bills recorded yet.[/dim]")
        return
    table = Table(
        title="Bill History",
        box=box.ROUNDED,
        border_style="yellow",
        title_style="bold yellow",
    )
    table.add_column("#", style="dim", justify="right")
    table.add_column("Payer", style="bold magenta")
    table.add_column("Amount", justify="right", style="green")
    table.add_column("Participants")
    for entry in splitter.history:
        table.add_row(
            str(entry["id"]),
            entry["payer"],
            f"{entry['amount']:.2f}",
            ", ".join(entry["consumers"]),
        )
    console.print(table)


def do_add_bill(splitter: DynamicBillSplitter):
    console.print("\n[bold cyan]--- Add a Bill ---[/bold cyan]")
    try:
        amount = FloatPrompt.ask("[bold]Amount[/bold]")
    except KeyboardInterrupt:
        return
    if amount <= 0:
        console.print("[bold red]Error:[/bold red] Amount must be positive.")
        return

    payer = Prompt.ask("[bold]Who paid?[/bold]").strip()
    if not payer:
        console.print("[bold red]Error:[/bold red] Payer name cannot be empty.")
        return

    people = splitter.known_people
    if payer not in people:
        people.append(payer)
    if people:
        console.print(f"[dim]Known group: {', '.join(people)}[/dim]")
    console.print("[dim]Enter consumer names separated by spaces/commas, or press Enter for everyone.[/dim]")

    cons_input = Prompt.ask("[bold]Consumers[/bold]", default="").strip()
    if not cons_input:
        consumers = list(people)
    else:
        consumers = [n.strip() for n in cons_input.replace(",", " ").split() if n.strip()]

    if not consumers:
        console.print("[bold red]Error:[/bold red] At least one consumer is required.")
        return

    splitter.add_bill(amount, payer, consumers)
    console.print(
        f"\n[green bold]Recorded:[/green bold] "
        f"[magenta]{payer}[/magenta] paid [green]{amount:.2f}[/green] "
        f"for [cyan]{', '.join(consumers)}[/cyan]"
    )

    console.print()
    show_balances(splitter)


def do_delete_bill(splitter: DynamicBillSplitter):
    if not splitter.history:
        console.print("[dim]No bills to delete.[/dim]")
        return

    console.print()
    show_history(splitter)
    console.print()

    try:
        bill_id_str = Prompt.ask("[bold]Enter bill # to delete[/bold]").strip()
        bill_id = int(bill_id_str)
    except (ValueError, KeyboardInterrupt):
        console.print("[bold red]Error:[/bold red] Please enter a valid bill number.")
        return

    if splitter.delete_bill(bill_id):
        console.print(f"[green]Bill #{bill_id} deleted successfully.[/green]")
    else:
        console.print(f"[bold red]Error:[/bold red] Bill #{bill_id} not found.")


def show_settlements(splitter: DynamicBillSplitter):
    console.print()
    if splitter.history:
        show_history(splitter)
        console.print()

    results = splitter.get_settlements()
    if not results:
        content = Text("Everything is square!", style="bold green")
    else:
        lines = Text()
        for debtor, creditor, amount in results:
            lines.append("💸  ")
            lines.append(debtor, style="bold red")
            lines.append(" -> ", style="dim")
            lines.append(creditor, style="bold green")
            lines.append(f" : {amount:.2f}\n", style="bold yellow")
        content = lines

    console.print(
        Panel(
            content,
            title="[bold]Final Settlements[/bold]",
            border_style="bright_green",
            box=box.DOUBLE,
            padding=(1, 2),
        )
    )
    out_path = write_session_output_file(splitter, results)
    console.print(f"[dim]Session saved to:[/dim] [cyan]{out_path.resolve()}[/cyan]")
    console.print("[bold]Done! Have a great day![/bold]\n")


def write_session_output_file(
    splitter: DynamicBillSplitter,
    settlements: list[tuple[str, str, float]],
    path: Path | None = None,
) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if path is None:
        path = Path(f"tomato_bill_session_{stamp}.xlsx")

    wb = Workbook()
    ws_bills = wb.active
    ws_bills.title = "Bills"
    ws_bills["A1"] = "Tomato Bill Splitter"
    ws_bills["B1"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"

    header_row = 3
    for col, title in enumerate(["#", "Payer", "Amount", "Participants"], start=1):
        ws_bills.cell(row=header_row, column=col, value=title)

    data_row = header_row + 1
    if not splitter.history:
        ws_bills.cell(row=data_row, column=1, value="(no bills recorded)")
    else:
        for entry in splitter.history:
            ws_bills.cell(row=data_row, column=1, value=entry["id"])
            ws_bills.cell(row=data_row, column=2, value=entry["payer"])
            ws_bills.cell(row=data_row, column=3, value=round(entry["amount"], 2))
            ws_bills.cell(row=data_row, column=4, value=", ".join(entry["consumers"]))
            data_row += 1

    ws_settle = wb.create_sheet("Settlements")
    ws_settle["A1"] = "From"
    ws_settle["B1"] = "To"
    ws_settle["C1"] = "Amount"
    if not settlements:
        ws_settle["A2"] = "Everything is square — no payments needed."
    else:
        r = 2
        for debtor, creditor, amount in settlements:
            ws_settle.cell(row=r, column=1, value=debtor)
            ws_settle.cell(row=r, column=2, value=creditor)
            ws_settle.cell(row=r, column=3, value=round(amount, 2))
            r += 1

    wb.save(path)
    return path


def main():
    show_banner()
    splitter = DynamicBillSplitter()

    while True:
        show_menu()
        choice = Prompt.ask("[bold cyan]Choose[/bold cyan]", choices=["1", "2", "3", "4", "5"], show_choices=False)

        if choice == "1":
            do_add_bill(splitter)
        elif choice == "2":
            console.print()
            show_balances(splitter)
        elif choice == "3":
            console.print()
            show_history(splitter)
        elif choice == "4":
            do_delete_bill(splitter)
        elif choice == "5":
            show_settlements(splitter)
            break


if __name__ == "__main__":
    main()
